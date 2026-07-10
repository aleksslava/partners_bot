import asyncio
import logging
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from aiogram import Bot
from aiogram.exceptions import TelegramRetryAfter
from aiogram.types import FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import load_workbook

from web_admin.max_client import (
    MaxBroadcastClient,
    MaxDeliveryError,
    MaxServiceUnavailable,
)
from web_admin.storage import BroadcastStorage

logger = logging.getLogger(__name__)

NAME_MASK = '[Имя]'
MAX_XLSX_SIZE = 10 * 1024 * 1024
MAX_MEDIA_SIZE = 20 * 1024 * 1024
TELEGRAM_TEXT_LIMIT = 4096
TELEGRAM_CAPTION_LIMIT = 1024
MAX_TEXT_LIMIT = 4000
ALLOWED_MEDIA = {
    '.jpg': 'photo',
    '.jpeg': 'photo',
    '.png': 'photo',
    '.mp4': 'video',
}
PLATFORM_COLUMNS = {
    'telegram': ('telegram_id', 'telegram_id', 'raw_telegram_id'),
    'max': ('max_id', 'max_id', 'raw_max_id'),
}


class UploadValidationError(ValueError):
    pass


def normalize_recipient_id(value: Any) -> int:
    if isinstance(value, bool):
        raise ValueError
    if isinstance(value, int):
        recipient_id = value
    elif isinstance(value, float) and value.is_integer():
        recipient_id = int(value)
    else:
        text = str(value or '').strip()
        if not text.isdigit():
            raise ValueError
        recipient_id = int(text)
    if recipient_id <= 0:
        raise ValueError
    return recipient_id


def normalize_telegram_id(value: Any) -> int:
    return normalize_recipient_id(value)


def message_limit(targets: set[str], has_media: bool) -> int:
    limits: list[int] = []
    if 'telegram' in targets:
        limits.append(TELEGRAM_CAPTION_LIMIT if has_media else TELEGRAM_TEXT_LIMIT)
    if 'max' in targets:
        limits.append(MAX_TEXT_LIMIT)
    if not limits:
        raise UploadValidationError('Выберите хотя бы одного бота.')
    return min(limits)


def parse_recipients(
    file_content: bytes,
    message: str,
    *,
    targets: set[str] | None = None,
    message_limit_value: int = TELEGRAM_TEXT_LIMIT,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, int]]]:
    targets = targets or {'telegram'}
    unknown_targets = targets.difference(PLATFORM_COLUMNS)
    if unknown_targets or not targets:
        raise UploadValidationError('Выберите хотя бы одного поддерживаемого бота.')
    if not file_content:
        raise UploadValidationError('Excel-файл пуст.')
    if len(file_content) > MAX_XLSX_SIZE:
        raise UploadValidationError('Excel-файл должен быть не больше 10 МБ.')

    try:
        workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as error:
        raise UploadValidationError('Не удалось открыть Excel-файл.') from error

    try:
        sheet = workbook.active
        iter_rows = getattr(sheet, 'iter_rows', None)
        if iter_rows is None:
            raise UploadValidationError('В Excel-файле нет активного листа.')
        header_cells = next(iter_rows(min_row=1, max_row=1), None)
        if header_cells is None:
            raise UploadValidationError('В Excel-файле нет заголовков.')
        columns = {
            str(cell.value).strip().casefold(): index
            for index, cell in enumerate(header_cells)
            if cell.value is not None
        }
        missing_columns = [
            column for column in ('telegram_id', 'max_id') if column not in columns
        ]
        if missing_columns:
            raise UploadValidationError(
                f"Не найдены обязательные колонки: {', '.join(missing_columns)}."
            )
        name_column = columns.get('имя')
        if name_column is None:
            name_column = columns.get('название')

        recipients: list[dict[str, Any]] = []
        seen_ids: dict[str, set[int]] = {platform: set() for platform in targets}
        stats = {
            platform: {'ready': 0, 'skipped': 0, 'duplicates': 0, 'invalid': 0}
            for platform in targets
        }
        needs_name = NAME_MASK in message

        for row_number, row in enumerate(iter_rows(min_row=2, values_only=True), start=2):
            if all(value in (None, '') for value in row):
                continue
            name_value = row[name_column] if name_column is not None and name_column < len(row) else None
            name = str(name_value or '').strip()
            recipient: dict[str, Any] = {
                'row_number': row_number,
                'name': name,
                'deliveries': [],
            }
            for platform, (header, id_key, raw_key) in PLATFORM_COLUMNS.items():
                raw_value = row[columns[header]] if columns[header] < len(row) else None
                recipient[raw_key] = str(raw_value or '').strip()
                try:
                    recipient[id_key] = normalize_recipient_id(raw_value)
                except ValueError:
                    recipient[id_key] = None

            for platform in targets:
                _, id_key, raw_key = PLATFORM_COLUMNS[platform]
                target_id = recipient.get(id_key)
                delivery = {
                    'platform': platform,
                    'target_id': target_id,
                    'raw_target_id': recipient.get(raw_key),
                    'status': 'pending',
                    'error': None,
                }
                if target_id is None:
                    delivery.update(status='skipped', error=f'Некорректный {id_key}')
                    stats[platform]['invalid'] += 1
                elif target_id in seen_ids[platform]:
                    delivery.update(status='skipped', error=f'Повторный {id_key}')
                    stats[platform]['duplicates'] += 1
                elif needs_name and not name:
                    delivery.update(status='skipped', error='Не указано имя для [Имя]')
                    stats[platform]['invalid'] += 1
                elif len(render_message(message, name)) > message_limit_value:
                    delivery.update(
                        status='skipped',
                        error='Сообщение после подстановки слишком длинное',
                    )
                    stats[platform]['invalid'] += 1
                else:
                    seen_ids[platform].add(target_id)
                    stats[platform]['ready'] += 1
                if delivery['status'] == 'skipped':
                    stats[platform]['skipped'] += 1
                recipient['deliveries'].append(delivery)
            recipients.append(recipient)
    finally:
        workbook.close()

    if not recipients:
        raise UploadValidationError('В Excel-файле нет получателей.')
    if not any(
        delivery['status'] == 'pending'
        for recipient in recipients
        for delivery in recipient['deliveries']
    ):
        raise UploadValidationError('В Excel-файле нет корректных получателей.')
    return recipients, stats


def render_message(message: str, name: str) -> str:
    return message.replace(NAME_MASK, name)


def build_button(text: str | None, url: str | None) -> InlineKeyboardMarkup | None:
    if not text or not url:
        return None
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=text, url=url)]],
    )


class BroadcastService:
    def __init__(
        self,
        storage: BroadcastStorage,
        bot: Bot,
        max_client: MaxBroadcastClient | None = None,
    ):
        self.storage = storage
        self.bot = bot
        self.max_client = max_client
        self.media_dir = storage.data_dir / 'media'
        self._worker_task: asyncio.Task | None = None
        self._wake_event = asyncio.Event()

    def initialize(self) -> None:
        self.storage.initialize()
        self.media_dir.mkdir(parents=True, exist_ok=True)
        self.storage.recover_interrupted()
        for media_path in self.storage.stale_draft_media():
            self.delete_media(media_path)

    def start(self) -> None:
        if self._worker_task is None or self._worker_task.done():
            self._worker_task = asyncio.create_task(self._worker_loop())

    async def stop(self) -> None:
        if self._worker_task is not None:
            self._worker_task.cancel()
            try:
                await self._worker_task
            except asyncio.CancelledError:
                pass
            self._worker_task = None
        if self.max_client is not None:
            await self.max_client.close()

    def wake(self) -> None:
        self._wake_event.set()

    async def _worker_loop(self) -> None:
        while True:
            processed = await self.process_next_due()
            if processed:
                continue
            self._wake_event.clear()
            try:
                await asyncio.wait_for(self._wake_event.wait(), timeout=1.0)
            except asyncio.TimeoutError:
                pass

    async def process_next_due(self) -> bool:
        broadcast = self.storage.next_due(datetime.now(timezone.utc))
        if broadcast is None:
            return False
        broadcast_id = int(broadcast['id'])
        if not self.storage.mark_running(broadcast_id):
            return True

        try:
            await self._prepare_max_media(broadcast)
            broadcast = self.storage.get_broadcast(broadcast_id) or broadcast
            deliveries = self.storage.pending_deliveries(broadcast_id)
            for delivery in deliveries:
                delivery_id = int(delivery['id'])
                if not self.storage.mark_delivery_sending(delivery_id):
                    continue
                try:
                    if delivery['platform'] == 'telegram':
                        await self._send_telegram(broadcast, delivery)
                    else:
                        await self._send_max(broadcast, delivery)
                except MaxServiceUnavailable as error:
                    self.storage.mark_delivery_result(
                        delivery_id,
                        success=False,
                        error=str(error)[:500],
                    )
                    self.storage.fail_pending_platform(
                        broadcast_id,
                        'max',
                        str(error)[:500],
                    )
                except Exception as error:
                    logger.exception(
                        'Ошибка веб-рассылки %s platform=%s target_id=%s',
                        broadcast_id,
                        delivery['platform'],
                        delivery['target_id'],
                    )
                    self.storage.mark_delivery_result(
                        delivery_id,
                        success=False,
                        error=str(error)[:500],
                    )
                else:
                    self.storage.mark_delivery_result(delivery_id, success=True)
                await asyncio.sleep(0.05)
            self.storage.finish_broadcast(broadcast_id)
        except Exception as error:
            logger.exception('Не удалось выполнить веб-рассылку %s', broadcast_id)
            self.storage.fail_broadcast(broadcast_id, str(error)[:500])
        finally:
            self.delete_media(broadcast.get('media_path'))
        return True

    async def _prepare_max_media(self, broadcast: dict[str, Any]) -> None:
        if not broadcast.get('send_max') or not broadcast.get('media_path'):
            return
        if broadcast.get('max_media_token'):
            return
        broadcast_id = int(broadcast['id'])
        if self.max_client is None:
            self.storage.fail_pending_platform(
                broadcast_id,
                'max',
                'Интеграция MAX не настроена',
            )
            return
        try:
            media = await self.max_client.upload_media(broadcast['media_path'])
        except (MaxServiceUnavailable, MaxDeliveryError) as error:
            self.storage.fail_pending_platform(broadcast_id, 'max', str(error)[:500])
            return
        self.storage.save_max_media(
            broadcast_id,
            media_type=media['media_type'],
            token=media['token'],
        )

    async def _send_telegram(
        self,
        broadcast: dict[str, Any],
        delivery: dict[str, Any],
    ) -> None:
        message = render_message(broadcast['message'], delivery.get('name', ''))
        keyboard = build_button(broadcast.get('button_text'), broadcast.get('button_url'))

        async def send() -> None:
            if broadcast.get('media_kind') == 'photo':
                await self.bot.send_photo(
                    chat_id=int(delivery['target_id']),
                    photo=FSInputFile(broadcast['media_path']),
                    caption=message,
                    reply_markup=keyboard,
                    parse_mode=None,
                )
            elif broadcast.get('media_kind') == 'video':
                await self.bot.send_video(
                    chat_id=int(delivery['target_id']),
                    video=FSInputFile(broadcast['media_path']),
                    caption=message,
                    reply_markup=keyboard,
                    parse_mode=None,
                    supports_streaming=True,
                )
            else:
                await self.bot.send_message(
                    chat_id=int(delivery['target_id']),
                    text=message,
                    reply_markup=keyboard,
                    parse_mode=None,
                )

        try:
            await send()
        except TelegramRetryAfter as error:
            await asyncio.sleep(float(error.retry_after))
            await send()

    async def _send_max(
        self,
        broadcast: dict[str, Any],
        delivery: dict[str, Any],
    ) -> None:
        if self.max_client is None:
            raise MaxServiceUnavailable('Интеграция MAX не настроена')
        await self.max_client.send_message(
            max_id=int(delivery['target_id']),
            text=render_message(broadcast['message'], delivery.get('name', '')),
            button_text=broadcast.get('button_text'),
            button_url=broadcast.get('button_url'),
            media_type=broadcast.get('max_media_type'),
            media_token=broadcast.get('max_media_token'),
        )

    @staticmethod
    def delete_media(media_path: str | None) -> None:
        if not media_path:
            return
        try:
            Path(media_path).unlink(missing_ok=True)
        except OSError:
            logger.exception('Не удалось удалить медиафайл %s', media_path)

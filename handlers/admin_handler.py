import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Final

from aiogram import Bot, F, Router
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from keybooards.main_keyboards import get_start_keyboard
from lexicon.lexicon_ru import (
    spam_message,
    spam_message_2,
    spam_url,
    spam_url_2,
    start_menu,
)

admin_router = Router()
logger = logging.getLogger(__name__)

CUSTOMERS_XLSX_PATH = Path('media/xls_files/customers.xlsx')
TRAINING_CUSTOMERS_XLSX_PATH = Path('media/xls_files/customers_2.xlsx')
SPAM_VIDEO_PATH = Path('media/video/video.mp4')
TRAINING_SPAM_VIDEO_PATH = Path('media/video/video_2.mp4')
SPAM_NAME_MASK = '[Имя]'
SPAM_STATUS_SUCCESS = 'Успешно'
SPAM_STATUS_ERROR = 'Ошибка отправки'
SPAM_STATUS_SKIP = 'Пропуск'
BROADCAST_KIND_INTERVIEW: Final = 'interview'
BROADCAST_KIND_TRAINING: Final = 'training'


@dataclass(frozen=True)
class BroadcastConfig:
    """Configuration for one admin broadcast type."""

    xlsx_path: Path
    video_path: Path
    message: str
    button_text: str
    button_url: str


BROADCAST_CONFIGS: Final[dict[str, BroadcastConfig]] = {
    BROADCAST_KIND_INTERVIEW: BroadcastConfig(
        xlsx_path=CUSTOMERS_XLSX_PATH,
        video_path=SPAM_VIDEO_PATH,
        message=spam_message,
        button_text='Записаться на интервью',
        button_url=spam_url,
    ),
    BROADCAST_KIND_TRAINING: BroadcastConfig(
        xlsx_path=TRAINING_CUSTOMERS_XLSX_PATH,
        video_path=TRAINING_SPAM_VIDEO_PATH,
        message=spam_message_2,
        button_text='Записаться на обучение',
        button_url=spam_url_2,
    ),
}


class AdminSpamStates(StatesGroup):
    waiting_range = State()
    waiting_single_tg_id = State()
    waiting_single_name = State()


def is_admin(user_id: int, admin_id: str) -> bool:
    try:
        return int(user_id) == int(admin_id)
    except (TypeError, ValueError):
        return False


def admin_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(
                text='Сделать рассылку',
                callback_data='admin_broadcast',
            ),
        ],
        [
            InlineKeyboardButton(
                text='Сделать рассылку на обучение',
                callback_data='admin_training_broadcast',
            ),
        ],
        [
            InlineKeyboardButton(
                text='Отправить пользователю по telegram_id',
                callback_data='admin_single',
            ),
        ],
        [
            InlineKeyboardButton(
                text='Назад в главное меню',
                callback_data='admin_back_main',
            ),
        ],
    ])


def spam_keyboard(config: BroadcastConfig) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=config.button_text, url=config.button_url)]
    ])


def spam_text(name: str, message: str = spam_message) -> str:
    return message.replace(SPAM_NAME_MASK, name.strip())


def parse_spam_range(text: str) -> tuple[int, int] | None:
    match = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', text or '')
    if not match:
        return None
    start, end = int(match.group(1)), int(match.group(2))
    if start < 1 or end < start:
        return None
    return start, end


def normalize_telegram_id(value) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    value = str(value).strip()
    if value.isdigit():
        return int(value)
    raise ValueError('Некорректный telegram_id')


def get_xlsx_columns(sheet: Worksheet) -> dict[str, int]:
    columns = {}
    for cell in sheet[1]:
        if cell.value is not None:
            columns[str(cell.value).strip()] = cell.column
    return columns


async def edit_callback_text(
    callback: CallbackQuery,
    text: str,
    reply_markup: InlineKeyboardMarkup | None = None,
) -> bool:
    if not isinstance(callback.message, Message):
        await callback.answer(text='Сообщение недоступно', show_alert=True)
        return False

    await callback.message.edit_text(text=text, reply_markup=reply_markup)
    return True


async def send_interview_spam(bot: Bot, chat_id: int, name: str) -> None:
    await send_admin_spam(
        bot=bot,
        chat_id=chat_id,
        name=name,
        config=BROADCAST_CONFIGS[BROADCAST_KIND_INTERVIEW],
    )


async def send_admin_spam(
    bot: Bot,
    chat_id: int,
    name: str,
    config: BroadcastConfig,
) -> None:
    await bot.send_video(
        chat_id=chat_id,
        video=FSInputFile(config.video_path),
        caption=spam_text(name=name, message=config.message),
        reply_markup=spam_keyboard(config),
        width=1080,
        height=1920,
        supports_streaming=True,
    )


@admin_router.callback_query(F.data == 'start_admin')
async def start_admin_menu(
    callback: CallbackQuery,
    admin_id: str,
    state: FSMContext,
) -> None:
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.clear()
    is_edited = await edit_callback_text(
        callback=callback,
        text='<b>Меню администратора</b>',
        reply_markup=admin_menu_keyboard(),
    )
    if not is_edited:
        return
    await callback.answer()


@admin_router.callback_query(F.data == 'admin_back_main')
async def admin_back_main(
    callback: CallbackQuery,
    admin_id: str,
    state: FSMContext,
) -> None:
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.clear()
    is_edited = await edit_callback_text(
        callback=callback,
        text='<b>Основное меню чат-бота HiTE PRO!</b>',
        reply_markup=await get_start_keyboard(start_menu, is_admin=True),
    )
    if not is_edited:
        return
    await callback.answer()


@admin_router.callback_query(F.data == 'admin_broadcast')
async def start_admin_broadcast(
    callback: CallbackQuery,
    admin_id: str,
    state: FSMContext,
) -> None:
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.update_data(broadcast_kind=BROADCAST_KIND_INTERVIEW)
    await state.set_state(AdminSpamStates.waiting_range)
    is_edited = await edit_callback_text(
        callback=callback,
        text='Введите диапазон пользователей для рассылки в формате N-N.',
    )
    if not is_edited:
        return
    await callback.answer()


@admin_router.callback_query(F.data == 'admin_training_broadcast')
async def start_admin_training_broadcast(
    callback: CallbackQuery,
    admin_id: str,
    state: FSMContext,
) -> None:
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.update_data(broadcast_kind=BROADCAST_KIND_TRAINING)
    await state.set_state(AdminSpamStates.waiting_range)
    is_edited = await edit_callback_text(
        callback=callback,
        text='Введите диапазон пользователей для рассылки в формате N-N.',
    )
    if not is_edited:
        return
    await callback.answer()


@admin_router.message(StateFilter(AdminSpamStates.waiting_range), F.text)
async def process_admin_broadcast_range(
    message: Message,
    bot: Bot,
    admin_id: str,
    state: FSMContext,
) -> None:
    if message.from_user is None or not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    spam_range = parse_spam_range(message.text or '')
    if spam_range is None:
        await message.answer(
            text='Некорректный диапазон. Введите диапазон в формате N-N, '
                 'например: 1-10.',
        )
        return

    state_data = await state.get_data()
    broadcast_kind = state_data.get('broadcast_kind', BROADCAST_KIND_INTERVIEW)
    if not isinstance(broadcast_kind, str):
        broadcast_kind = BROADCAST_KIND_INTERVIEW
    config = BROADCAST_CONFIGS.get(
        broadcast_kind,
        BROADCAST_CONFIGS[BROADCAST_KIND_INTERVIEW],
    )

    if not config.xlsx_path.exists():
        await state.clear()
        await message.answer(text=f'Файл {config.xlsx_path} не найден.')
        return

    if not config.video_path.exists():
        await state.clear()
        await message.answer(text=f'Файл {config.video_path} не найден.')
        return

    try:
        workbook = load_workbook(config.xlsx_path)
    except Exception as error:
        logger.exception('Не удалось открыть Excel-файл для рассылки')
        await state.clear()
        await message.answer(text=f'Не удалось открыть Excel-файл: {error}')
        return

    active_sheet = workbook.active
    if not isinstance(active_sheet, Worksheet):
        workbook.close()
        await state.clear()
        await message.answer(text='В Excel-файле не найден активный лист.')
        return

    sheet = active_sheet
    columns = get_xlsx_columns(sheet)
    required_columns = ['Название', 'telegram_id', 'Отправлено в ТГ']
    missed_columns = [column for column in required_columns if column not in columns]
    if missed_columns:
        workbook.close()
        await state.clear()
        missed_columns_text = ', '.join(missed_columns)
        await message.answer(
            text=f'В Excel-файле не найдены колонки: {missed_columns_text}.',
        )
        return

    start_client, end_client = spam_range
    max_client_number = max(sheet.max_row - 1, 0)
    if end_client > max_client_number:
        workbook.close()
        await state.clear()
        await message.answer(
            text=f'В Excel-файле только {max_client_number} строк с клиентами.',
        )
        return

    await message.answer(text='Рассылка запущена.')

    total = 0
    success = 0
    errors = 0
    name_column = columns['Название']
    telegram_id_column = columns['telegram_id']
    status_column = columns['Отправлено в ТГ']

    for row_number in range(start_client + 1, end_client + 2):
        telegram_id_cell = sheet.cell(row=row_number, column=telegram_id_column)
        status_cell = sheet.cell(row=row_number, column=status_column)

        if telegram_id_cell.value is None or str(telegram_id_cell.value).strip() == '':
            status_cell.value = SPAM_STATUS_SKIP
            continue

        total += 1
        full_name = str(
            sheet.cell(row=row_number, column=name_column).value or '',
        ).strip()
        name = full_name.split(maxsplit=1)[0] if full_name else ''

        try:
            telegram_id = normalize_telegram_id(telegram_id_cell.value)
            await send_admin_spam(
                bot=bot,
                chat_id=telegram_id,
                name=name,
                config=config,
            )
            status_cell.value = SPAM_STATUS_SUCCESS
            success += 1
        except Exception:
            status_cell.value = SPAM_STATUS_ERROR
            errors += 1
            logger.exception(
                'Ошибка отправки рассылки пользователю %s',
                telegram_id_cell.value,
            )

    save_error = None
    try:
        workbook.save(config.xlsx_path)
    except Exception as error:
        save_error = error
        logger.exception('Не удалось сохранить Excel-файл после рассылки')
    finally:
        workbook.close()

    await state.clear()
    stats_message = (f'Всего отправок: {total}\n'
                     f'Успешных: {success}\n'
                     f'Отправок с ошибкой: {errors}')
    if save_error is not None:
        stats_message += f'\n\nНе удалось сохранить Excel-файл: {save_error}'
    await message.answer(text=stats_message)


@admin_router.callback_query(F.data == 'admin_single')
async def start_admin_single_send(
    callback: CallbackQuery,
    admin_id: str,
    state: FSMContext,
) -> None:
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.set_state(AdminSpamStates.waiting_single_tg_id)
    is_edited = await edit_callback_text(
        callback=callback,
        text='Введите telegram_id пользователя.',
    )
    if not is_edited:
        return
    await callback.answer()


@admin_router.message(StateFilter(AdminSpamStates.waiting_single_tg_id), F.text)
async def process_admin_single_tg_id(
    message: Message,
    admin_id: str,
    state: FSMContext,
) -> None:
    if message.from_user is None or not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    try:
        telegram_id = normalize_telegram_id(message.text or '')
    except ValueError:
        await message.answer(
            text='Некорректный telegram_id. Введите только числовой telegram_id.',
        )
        return

    await state.update_data(telegram_id=telegram_id)
    await state.set_state(AdminSpamStates.waiting_single_name)
    await message.answer(text='Введите имя получателя для подстановки в сообщение.')


@admin_router.message(StateFilter(AdminSpamStates.waiting_single_name), F.text)
async def process_admin_single_name(
    message: Message,
    bot: Bot,
    admin_id: str,
    state: FSMContext,
) -> None:
    if message.from_user is None or not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    if not SPAM_VIDEO_PATH.exists():
        await state.clear()
        await message.answer(text=f'Файл {SPAM_VIDEO_PATH} не найден.')
        return

    state_data = await state.get_data()
    telegram_id = state_data.get('telegram_id')
    if not isinstance(telegram_id, int):
        await state.clear()
        await message.answer(text='telegram_id не найден в состоянии рассылки.')
        return

    name = (message.text or '').strip()

    try:
        await send_interview_spam(bot=bot, chat_id=telegram_id, name=name)
    except Exception as error:
        logger.exception('Ошибка отправки сообщения пользователю %s', telegram_id)
        await state.clear()
        await message.answer(text=f'Ошибка отправки: {error}')
        return

    await state.clear()
    await message.answer(text='Сообщение успешно отправлено.')

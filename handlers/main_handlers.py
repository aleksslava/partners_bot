import logging
import json
import os
import re
from pathlib import Path

import aiogram.exceptions
from openpyxl import load_workbook
from redis.asyncio.client import Redis
from service.service import LeadData
from service.service import Order, get_kp_pdf

from aiogram import Router, F, Bot
from aiogram.filters import Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import Message, CallbackQuery, WebAppInfo, \
    KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup

from keybooards.main_keyboards import (reply_phone_number, get_contacts_list, hide_contacts_list, get_start_keyboard,
                                       forum_button, manager_button, support_button, problem_button,
                                       helpfull_materials_keyboard, back_button, answer_for_user, authorized_client,
                                       link_to_opt_button, confirm_spam,
                                       )
from config_data.amo_api import AmoCRMWrapper, Contact
from lexicon.lexicon_ru import account_info, Lexicon_RU, start_menu, helpfull_materials_menu, new_shop_inline_message, \
    spam_message, spam_url

main_router = Router()
logger = logging.getLogger(__name__)

CUSTOMERS_XLSX_PATH = Path('media/xls_files/customers.xlsx')
SPAM_VIDEO_PATH = Path('media/video/video.mp4')
SPAM_NAME_MASK = '[Имя]'
SPAM_STATUS_SUCCESS = 'Успешно'
SPAM_STATUS_ERROR = 'Ошибка отправки'
SPAM_STATUS_SKIP = 'Пропуск'


class AdminSpamStates(StatesGroup):
    waiting_range = State()
    waiting_single_tg_id = State()
    waiting_single_name = State()


def is_admin(user_id: int, admin_id: str) -> bool:
    try:
        return int(user_id) == int(admin_id)
    except (TypeError, ValueError):
        return False


def admin_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Сделать рассылку', callback_data='admin_broadcast')],
        [InlineKeyboardButton(text='Отправить пользователю по telegram_id', callback_data='admin_single')],
        [InlineKeyboardButton(text='Назад в главное меню', callback_data='/start')],
    ])


def spam_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Записаться на интервью', url=spam_url)]
    ])


def spam_text(name: str) -> str:
    return spam_message.replace(SPAM_NAME_MASK, name.strip())


def parse_spam_range(text: str) -> tuple[int, int] | None:
    match = re.fullmatch(r'\s*(\d+)\s*-\s*(\d+)\s*', text or '')
    if not match:
        return None
    start, end = int(match.group(1)), int(match.group(2))
    if start < 1 or end < start:
        return None
    return start, end


def normalize_telegram_id(value) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    value = str(value).strip()
    if value.isdigit():
        return int(value)
    raise ValueError('Некорректный telegram_id')


def get_xlsx_columns(sheet) -> dict[str, int]:
    columns = {}
    for cell in sheet[1]:
        if cell.value is not None:
            columns[str(cell.value).strip()] = cell.column
    return columns


async def send_interview_spam(bot: Bot, chat_id: int, name: str) -> None:
    await bot.send_video(
        chat_id=chat_id,
        video=FSInputFile(SPAM_VIDEO_PATH),
        caption=spam_text(name),
        reply_markup=spam_keyboard()
    )


@main_router.message(CommandStart())  # Хэндлер для обработки команды /start
async def command_start_process(message: Message, admin_id: str):
    tg_user_id = message.from_user.id
    if tg_user_id == int(admin_id):
        await message.answer(text='<b>Основное меню чат-бота HiTE PRO!</b>',
                             reply_markup=await get_start_keyboard(start_menu, is_admin=True))
    else:
        await message.answer(text='<b>Основное меню чат-бота HiTE PRO!</b>',
                             reply_markup=await get_start_keyboard(start_menu))

@main_router.callback_query(F.data == '/start')
async def command_start_inline_process(callback: CallbackQuery, admin_id: str, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(text='<b>Основное меню чат-бота HiTE PRO!</b>',
                                     reply_markup=await get_start_keyboard(start_menu,
                                                                           is_admin=is_admin(callback.from_user.id,
                                                                                             admin_id)))

@main_router.message(Command(commands=['info']))  # Хэндлер для обработки команды /info
async def info_handler(message: Message, amo_api: AmoCRMWrapper, fields_id: dict, bot: Bot):
    tg_id = message.from_user.id

    # Проверка наличия партнёра в бд по tg_id

    customer = amo_api.get_customer_by_tg_id(tg_id)
    if customer.get('status_code'):
        if customer.get('tg_id_in_db'):
            customer = customer.get('response')
            responsible_manager = amo_api.get_responsible_user_by_id(int(customer.get('responsible_user_id')))
            customer['manager'] = responsible_manager
            customer_params = amo_api.get_customer_params(customer, fields_id=fields_id)
            await message.answer(text=account_info(customer_params),
                                 reply_markup=await get_contacts_list(customer_params.id))
        else:
            # Если tg_id нет в бд, то ищем по номеру телефона
            name = message.from_user.first_name
            # await message.answer(text=f'{name}, здравствуйте.\n'
            #                           f'Поделитесь своим номером телефона для использования бота.👇',
            #                      reply_markup=await reply_phone_number())
            file = FSInputFile('image.png')
            await bot.send_photo(chat_id=message.chat.id, photo=file,
                                 caption=f'{name}, здравствуйте.\n'
                                         f'Поделитесь своим номером телефона для использования бота.👇',
                                 show_caption_above_media=True,
                                 reply_markup=await reply_phone_number())
    else:
        await message.answer(text='Ошибка! Помогите нам её исправить. Сообщите об этой ошибке в онлайн-форме:',
                             reply_markup=await problem_button())


@main_router.callback_query(F.data == '/info')  # Обработка инлайн кнопки "Мой профиль"
async def info_handler_cl(callback: CallbackQuery, amo_api: AmoCRMWrapper, fields_id: dict, bot: Bot):
    await callback.answer()
    tg_id = callback.message.chat.id

    # Проверка наличия партнёра в амо по tg_id

    customer = amo_api.get_customer_by_tg_id(tg_id)
    if customer.get('status_code'):  # Проверка корректности ответа от амо
        if customer.get('tg_id_in_db'):  # Проверка наличия tg_id в базе амо
            customer = customer.get('response')
            responsible_manager = amo_api.get_responsible_user_by_id(int(customer.get('responsible_user_id')))
            customer['manager'] = responsible_manager
            customer_params = amo_api.get_customer_params(customer, fields_id=fields_id)

            await callback.message.edit_text(text=account_info(customer_params),
                                             reply_markup=await get_contacts_list(customer_params.id))

        else:
            # Если tg_id нет в бд, то ищем по номеру телефона
            name = callback.message.chat.first_name
            # await callback.message.answer(text=f'{name}, здравствуйте.\n'
            #                                    f'Поделитесь своим номером телефона для использования бота.👇',
            #                               reply_markup=await reply_phone_number())
            file = FSInputFile('image.png')
            await bot.send_photo(chat_id=callback.message.chat.id, photo=file,
                                 caption=f'{name}, здравствуйте.\n'
                                         f'Поделитесь своим номером телефона для использования бота.👇',
                                 show_caption_above_media=True,
                                 reply_markup=await reply_phone_number())
    else:
        await callback.message.edit_text(text='Ошибка! Помогите нам её исправить. '
                                              '👇 Сообщите об этой ошибке в онлайн-форме.',
                                         reply_markup=await problem_button())


@main_router.message(F.contact)  # Хэндлер для обработки отправленного пользователем контакта
async def get_contact(message: Message, amo_api: AmoCRMWrapper, fields_id: dict):
    contact_phone = message.contact.phone_number
    contact_username = '@' + message.from_user.username if message.from_user.username is not None else ''
    logger.info(f'Получен контакт клиента для авторизации:\n'
                f'Номер телефона: {contact_phone}\n'
                f'Username: {contact_username}\n'
                f'TG_ID: {message.from_user.id}')
    customer = amo_api.get_customer_by_phone(contact_phone)

    if customer[0]:
        contact_id = customer[2].get('id')
        responsible_manager = amo_api.get_responsible_user_by_id(int(customer[1].get('responsible_user_id')))
        customer[1]['manager'] = responsible_manager
        customer_params = amo_api.get_customer_params(customer[1], fields_id=fields_id)
        amo_api.put_tg_id_to_customer(customer_params.id, message.from_user.id)
        amo_api.put_tgid_username_to_contact(id_contact=contact_id,
                                             username=contact_username,
                                             tg_id=message.from_user.id,
                                             fields_id=fields_id.get('contacts_fields_id'))

        await message.answer(text=f'Вы успешно авторизовались в чат боте HiTE PRO!\n\n'
                                  f'Можете посмотреть информацию из Вашего профиля и воспользоваться '
                                  f'магазином HiTE PRO👇', reply_markup=await authorized_client(start_menu))
    else:
        await message.answer(text=f'{customer[1]}\n\n'
                                  f'Воспользоваться чат-ботом могут только авторизованные партнёры.\n'
                                  f'👇 Если вы действующий партнёр компании HiTE PRO, '
                                  f'сообщите об этой ошибке в онлайн-форме.',
                             reply_markup=await problem_button())


# Хэндлер для обработки инлайн кнопки "Показать контакты"
@main_router.callback_query(F.data.startswith('contacts_list'))
async def open_contacts_list(callback: CallbackQuery, amo_api: AmoCRMWrapper):
    last_message = callback.message.text
    customer_id = callback.data.split('_')[2]
    customer = amo_api.get_customer_by_id(customer_id, with_contacts=True)
    contacts_list_id = [contact.get('id') for contact in customer[1]['_embedded']['contacts']]
    last_message = last_message + '\n\n<b>Привязанные контакты к профилю</b>\n'

    for contact_id in contacts_list_id:
        contact_data = Contact(**amo_api.get_contact_by_id(contact_id))
        last_message = last_message + str(contact_data)

    await callback.message.edit_text(text=last_message, reply_markup=await hide_contacts_list(customer_id))


# Хэндлер для обработки инлайн кнопки "Скрыть контакты"
@main_router.callback_query(F.data.startswith('hide_contacts_list'))
async def hide_contact_list(callback: CallbackQuery):
    customer_id = callback.data.split('_')[3]
    last_text = callback.message.text
    hide_index = last_text.find('Привязанные')
    new_text = last_text[:hide_index]
    await callback.message.edit_text(text=new_text, reply_markup=await get_contacts_list(customer_id))


@main_router.message(Command(commands='contacts'))  # Хэндлер для обработки команды /contacts
async def contacts(message: Message):
    await message.answer(text=Lexicon_RU.get('contact_message'))


@main_router.callback_query(F.data == '/contacts')  # Хэндлер для обработки inline кнопки "contacts"
async def command_contacts_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('contact_message'))


@main_router.message(Command(commands='shop'))  # Хэндлер для обработки команды /shop
async def command_shop_process(message: Message, amo_api: AmoCRMWrapper, fields_id: dict, bot: Bot,
                               redis: Redis):
    tg_id = message.from_user.id
    user_name = message.from_user.username
    try:
        await redis.set(name=str(tg_id), value=user_name)
    except:
        logger.error(f'Не получиось записать в бд id: {tg_id}')
    customer = amo_api.get_customer_by_tg_id(tg_id)
    contact = amo_api.get_contact_by_tg_id(tg_id, fields_id=fields_id.get('contacts_fields_id'))

    if customer.get('status_code') and contact.get('status_code'):  # Проверка корректности ответа от амо
        if customer.get('tg_id_in_db') and contact.get('tg_id_in_db'):  # Проверка наличия tg_id в базе амо
            customer = customer.get('response')
            customer['manager'] = {'name': None}
            customer_params = amo_api.get_customer_params(customer, fields_id=fields_id)
            bonus = str(customer_params.bonuses).replace(' ', '')
            discont = ''.join(list(filter(lambda x: x.isdigit(), customer_params.status)))
            web_app_url = fields_id.get('web_app_url')
            contact = contact.get('response')
            contact_id = contact.get('id')

            kb_1 = KeyboardButton(text='Открыть магазин',
                                  web_app=WebAppInfo(
                                      url=f'{web_app_url}?bonus={bonus}&'
                                          f'id={contact_id}&discont={discont}'))
            webapp_keyboard_1 = ReplyKeyboardMarkup(is_persistent=True, keyboard=[[kb_1, ]],
                                                    resize_keyboard=True, one_time_keyboard=True)
            await bot.send_message(chat_id=message.chat.id, text=Lexicon_RU['bonus_message'])
            await message.answer(text='Для перехода в магазин воспользуйтесь кнопкой клавиатуры👇',
                                 reply_markup=webapp_keyboard_1)
        else:
            # Если tg_id нет в бд, то ищем по номеру телефона
            name = message.from_user.first_name
            file = FSInputFile('image.png')
            await bot.send_photo(chat_id=message.chat.id, photo=file,
                                 caption=f'{name}, здравствуйте.\n'
                                         f'Поделитесь своим номером телефона для использования бота.👇',
                                 show_caption_above_media=True,
                                 reply_markup=await reply_phone_number())
    else:
        if customer.get('status_code'):
            response = contact.get('response')
        else:
            response = customer.get('response')
        await message.answer(text=f'{response}\n\n'
                                  f'👇 Сообщите об этой ошибке в онлайн-форме.',
                             reply_markup=await problem_button())

@main_router.message(Command(commands='new_shop'))
async def new_shop(message: Message, bot: Bot):
    shop_button = InlineKeyboardButton(text='Открыть магазин', web_app=WebAppInfo(url='https://profi-shop.hite-pro.ru/telegram/'))
    main_menu = InlineKeyboardButton(text='В главное меню', callback_data='/start')
    webapp_keyboard_1 = InlineKeyboardMarkup(inline_keyboard=[[shop_button], [main_menu]],)
    await bot.send_message(chat_id=message.chat.id,  text=Lexicon_RU['bonus_message'], reply_markup=ReplyKeyboardRemove())
    await message.answer(text=new_shop_inline_message, reply_markup=webapp_keyboard_1)

@main_router.callback_query(F.data == '/new_shop')  # Хэндлер для обработки inline кнопки "new_shop"
async def command_new_shop_process(callback: CallbackQuery, bot: Bot):
    shop_button = InlineKeyboardButton(text='Открыть магазин',
                                       web_app=WebAppInfo(url='https://profi-shop.hite-pro.ru/telegram/'))
    main_menu = InlineKeyboardButton(text='В главное меню', callback_data='/start')
    webapp_keyboard_1 = InlineKeyboardMarkup(inline_keyboard=[[shop_button], [main_menu]], )
    await bot.send_message(chat_id=callback.message.chat.id, text=Lexicon_RU['bonus_message'],
                           reply_markup=ReplyKeyboardRemove())
    await callback.message.answer(text=new_shop_inline_message,
                         reply_markup=webapp_keyboard_1)


@main_router.callback_query(F.data == '/shop')  # Хэндлер для обработки inline кнопки "shop"
async def command_shop_process_cl(callback: CallbackQuery, amo_api: AmoCRMWrapper, fields_id: dict, bot: Bot,
                                  redis: Redis):
    tg_id = callback.from_user.id
    user_name = callback.from_user.username if callback.from_user.username is not None else ''
    try:
        await redis.set(name=str(tg_id), value=user_name)
    except BaseException as error:
        logger.error(error)
        logger.error(f'Не получилось записать в бд id: {tg_id}')
    customer = amo_api.get_customer_by_tg_id(tg_id)
    contact = amo_api.get_contact_by_tg_id(tg_id, fields_id=fields_id.get('contacts_fields_id'))

    if customer.get('status_code') and contact.get('status_code'):  # Проверка корректности ответа от амо
        if customer.get('tg_id_in_db') and contact.get('tg_id_in_db'):  # Проверка наличия tg_id в базе амо
            customer = customer.get('response')
            customer['manager'] = {'name': None}
            customer_params = amo_api.get_customer_params(customer, fields_id=fields_id)

            bonus = str(customer_params.bonuses).replace(' ', '')
            discont = ''.join(list(filter(lambda x: x.isdigit(), customer_params.status)))
            web_app_url = fields_id.get('web_app_url')
            contact = contact.get('response')
            contact_id = contact.get('id')

            kb_1 = KeyboardButton(text='Открыть магазин',
                                  web_app=WebAppInfo(
                                      url=f'{web_app_url}?bonus={bonus}&'
                                          f'id={contact_id}&discont={discont}'))
            webapp_keyboard_1 = ReplyKeyboardMarkup(keyboard=[[kb_1, ]], resize_keyboard=True)
            await bot.send_message(chat_id=callback.message.chat.id, text=Lexicon_RU['bonus_message'])
            await callback.message.answer(text='Для перехода в магазин воспользуйтесь кнопкой клавиатуры👇',
                                          reply_markup=webapp_keyboard_1)
        else:
            # Если tg_id нет в бд, то ищем по номеру телефона
            name = callback.from_user.first_name
            file = FSInputFile('image.png')
            await bot.send_photo(chat_id=callback.message.chat.id, photo=file,
                                 caption=f'{name}, здравствуйте.\n'
                                         f'Поделитесь своим номером телефона для использования бота.👇',
                                 show_caption_above_media=True,
                                 reply_markup=await reply_phone_number())
    else:
        if customer.get('status_code'):
            response = contact.get('response')
        else:
            response = customer.get('response')
        await callback.message.answer(text=f'{response}\n\n'
                                           f'👇 Сообщите об этой ошибке в онлайн-форме.',
                                      reply_markup=await problem_button())

    await callback.answer()


@main_router.message(Command(commands='forum'))  # Хэндлер для обработки команды /forum
async def command_forum_process(message: Message):
    await message.answer(text=Lexicon_RU.get('forum_message'), reply_markup=await forum_button())


@main_router.callback_query(F.data == '/forum')  # Хэндлер для обработки inline кнопки "forum"
async def command_forum_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('forum_message'), reply_markup=await forum_button())


@main_router.message(Command(commands='materials'))  # Хэндлер для обработки команды /materials
async def command_materials_process(message: Message):
    await message.answer(text='<b>Полезные материалы HiTE PRO.</b>\n\n'
                              '👇 Используйте кнопки ниже, чтобы выбрать раздел.',
                         reply_markup=await helpfull_materials_keyboard(helpfull_materials_menu))


@main_router.callback_query(F.data == '/materials')  # Хэндлер для обработки inline кнопки "materials"
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text='<b>Полезные материалы HiTE PRO.</b>\n\n'
                                          '👇 Используйте кнопки ниже, чтобы выбрать раздел.',
                                     reply_markup=await helpfull_materials_keyboard(helpfull_materials_menu))


@main_router.callback_query(F.data == 'first_message')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('helpful_materials').get('first_message'),
                                     reply_markup=await back_button())


@main_router.callback_query(F.data == 'second_message')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('helpful_materials').get('second_message'),
                                     reply_markup=await back_button())


@main_router.callback_query(F.data == 'third_message')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('helpful_materials').get('third_message'),
                                     reply_markup=await back_button())


@main_router.callback_query(F.data == 'forth_message')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('helpful_materials').get('forth_message'),
                                     reply_markup=await back_button())


@main_router.callback_query(F.data == 'five_message')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('helpful_materials').get('five_message'),
                                     reply_markup=await back_button())


@main_router.callback_query(F.data == 'back')
async def command_materials_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text='<b>Полезные материалы HiTE PRO.</b>\n\n'
                                          '👇 Используйте кнопки ниже, чтобы выбрать раздел.',
                                     reply_markup=await helpfull_materials_keyboard(helpfull_materials_menu))


@main_router.message(Command(commands='partners'))  # Хэндлер для обработки команды /partners
async def command_partners_process(message: Message):
    await message.answer(text=Lexicon_RU.get('partner_kanal'))


@main_router.callback_query(F.data == '/partners')  # Хэндлер для обработки inline кнопки "partners"
async def command_forum_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('partner_kanal'))


@main_router.message(Command(commands='manager'))  # Хэндлер для обработки команды /manager
async def command_manager_process(message: Message):
    await message.answer(text=Lexicon_RU.get('manager'), reply_markup=await manager_button())


@main_router.callback_query(F.data == '/manager')  # Хэндлер для обработки inline кнопки "manager"
async def command_manager_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('manager'), reply_markup=await manager_button())


@main_router.message(Command(commands='support'))  # Хэндлер для обработки команды /support
async def command_support_process(message: Message):
    await message.answer(text=Lexicon_RU.get('support'), reply_markup=await support_button())


@main_router.callback_query(F.data == '/support')  # Хэндлер для обработки inline кнопки "support"
async def command_support_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('support'), reply_markup=await support_button())


@main_router.message(Command(commands='problem'))  # Хэндлер для обработки команды /problem
async def command_problem_process(message: Message):
    await message.answer(text=Lexicon_RU.get('problem'), reply_markup=await problem_button())


@main_router.callback_query(F.data == '/problem')  # Хэндлер для обработки inline кнопки "problem"
async def command_problem_process_cl(callback: CallbackQuery):
    await callback.message.edit_text(text=Lexicon_RU.get('problem'), reply_markup=await problem_button())

@main_router.message(Command(commands='bot_instr'))
async def bot_instr(message: Message):
    await message.answer_video(video='BAACAgIAAxkBAAIEzGkDLp9OWJMOfxossVWEHioSDdtQAALjiwACJ0kZSB1oJGqY-v-vNgQ')

@main_router.callback_query(F.data == '/bot_instr')
async def bot_instr_cl(callback: CallbackQuery):
    await callback.message.answer_video(video='BAACAgIAAxkBAAIEzGkDLp9OWJMOfxossVWEHioSDdtQAALjiwACJ0kZSB1oJGqY-v-vNgQ')


@main_router.callback_query(F.data == 'start_admin')
async def start_admin_menu(callback: CallbackQuery, admin_id: str, state: FSMContext):
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.clear()
    await callback.message.edit_text(text='<b>Меню администратора</b>', reply_markup=admin_menu_keyboard())
    await callback.answer()


@main_router.callback_query(F.data == 'admin_broadcast')
async def start_admin_broadcast(callback: CallbackQuery, admin_id: str, state: FSMContext):
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.set_state(AdminSpamStates.waiting_range)
    await callback.message.edit_text(text='Введите диапазон пользователей для рассылки в формате N-N.')
    await callback.answer()


@main_router.message(StateFilter(AdminSpamStates.waiting_range), F.text)
async def process_admin_broadcast_range(message: Message, bot: Bot, admin_id: str, state: FSMContext):
    if not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    spam_range = parse_spam_range(message.text)
    if spam_range is None:
        await message.answer(text='Некорректный диапазон. Введите диапазон в формате N-N, например: 1-10.')
        return

    if not CUSTOMERS_XLSX_PATH.exists():
        await state.clear()
        await message.answer(text=f'Файл {CUSTOMERS_XLSX_PATH} не найден.')
        return

    if not SPAM_VIDEO_PATH.exists():
        await state.clear()
        await message.answer(text=f'Файл {SPAM_VIDEO_PATH} не найден.')
        return

    try:
        workbook = load_workbook(CUSTOMERS_XLSX_PATH)
    except BaseException as error:
        logger.exception('Не удалось открыть Excel-файл для рассылки')
        await state.clear()
        await message.answer(text=f'Не удалось открыть Excel-файл: {error}')
        return

    sheet = workbook.active
    columns = get_xlsx_columns(sheet)
    required_columns = ['Название', 'telegram_id', 'Отправлено в ТГ']
    missed_columns = [column for column in required_columns if column not in columns]
    if missed_columns:
        await state.clear()
        await message.answer(text=f'В Excel-файле не найдены колонки: {", ".join(missed_columns)}.')
        return

    start_client, end_client = spam_range
    max_client_number = max(sheet.max_row - 1, 0)
    if end_client > max_client_number:
        await state.clear()
        await message.answer(text=f'В Excel-файле только {max_client_number} строк с клиентами.')
        return

    await message.answer(text='Рассылка запущена.')

    total = 0
    success = 0
    errors = 0
    name_column = columns['Название']
    telegram_id_column = columns['telegram_id']
    status_column = columns['Отправлено в ТГ']

    for row_number in range(start_client + 1, end_client + 2):
        telegram_id_cell = sheet.cell(row=row_number, column=telegram_id_column)
        status_cell = sheet.cell(row=row_number, column=status_column)

        if telegram_id_cell.value is None or str(telegram_id_cell.value).strip() == '':
            status_cell.value = SPAM_STATUS_SKIP
            continue

        total += 1
        name = str(sheet.cell(row=row_number, column=name_column).value or '').strip()

        try:
            telegram_id = normalize_telegram_id(telegram_id_cell.value)
            await send_interview_spam(bot=bot, chat_id=telegram_id, name=name)
            status_cell.value = SPAM_STATUS_SUCCESS
            success += 1
        except BaseException as error:
            status_cell.value = SPAM_STATUS_ERROR
            errors += 1
            logger.exception('Ошибка отправки рассылки пользователю %s', telegram_id_cell.value)

    save_error = None
    try:
        workbook.save(CUSTOMERS_XLSX_PATH)
    except BaseException as error:
        save_error = error
        logger.exception('Не удалось сохранить Excel-файл после рассылки')

    await state.clear()
    stats_message = (f'Всего отправок: {total}\n'
                     f'Успешных: {success}\n'
                     f'Отправок с ошибкой: {errors}')
    if save_error is not None:
        stats_message += f'\n\nНе удалось сохранить Excel-файл: {save_error}'
    await message.answer(text=stats_message)


@main_router.callback_query(F.data == 'admin_single')
async def start_admin_single_send(callback: CallbackQuery, admin_id: str, state: FSMContext):
    if not is_admin(callback.from_user.id, admin_id):
        await callback.answer(text='Нет доступа', show_alert=True)
        return

    await state.set_state(AdminSpamStates.waiting_single_tg_id)
    await callback.message.edit_text(text='Введите telegram_id пользователя.')
    await callback.answer()


@main_router.message(StateFilter(AdminSpamStates.waiting_single_tg_id), F.text)
async def process_admin_single_tg_id(message: Message, admin_id: str, state: FSMContext):
    if not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    try:
        telegram_id = normalize_telegram_id(message.text)
    except ValueError:
        await message.answer(text='Некорректный telegram_id. Введите только числовой telegram_id.')
        return

    await state.update_data(telegram_id=telegram_id)
    await state.set_state(AdminSpamStates.waiting_single_name)
    await message.answer(text='Введите имя получателя для подстановки в сообщение.')


@main_router.message(StateFilter(AdminSpamStates.waiting_single_name), F.text)
async def process_admin_single_name(message: Message, bot: Bot, admin_id: str, state: FSMContext):
    if not is_admin(message.from_user.id, admin_id):
        await state.clear()
        await message.answer(text='Нет доступа')
        return

    if not SPAM_VIDEO_PATH.exists():
        await state.clear()
        await message.answer(text=f'Файл {SPAM_VIDEO_PATH} не найден.')
        return

    state_data = await state.get_data()
    telegram_id = state_data.get('telegram_id')
    name = message.text.strip()

    try:
        await send_interview_spam(bot=bot, chat_id=telegram_id, name=name)
    except BaseException as error:
        logger.exception('Ошибка отправки сообщения пользователю %s', telegram_id)
        await state.clear()
        await message.answer(text=f'Ошибка отправки: {error}')
        return

    await state.clear()
    await message.answer(text='Сообщение успешно отправлено.')


@main_router.message(lambda message: message.chat.id == -1003159184418) # Обработка сообщений из группы рассылки
async def message_from_group(message: Message):
    message_id = message.message_id
    await message.reply(text='!!!ВНИМАНИЕ!!!\nВы отправите рассылку по партнёрам компании HiTE PRO!\n\nВы уверены?',
                        reply_markup=await confirm_spam(message_id))


@main_router.callback_query(F.data.startswith('spam'))
async def forward_spam_message(callback: CallbackQuery, bot: Bot, redis: Redis):
    from_chat_id = -1003159184418
    answer, message_id = callback.data.split('_')
    if answer == 'spamyes':
        await callback.message.edit_text(text='Рассылка запущена.\nО результате будет оповещение.')
        good_try = 0
        bad_try = 0
        async for chat_id in redis.scan_iter():
            try:
                await bot.copy_message(chat_id=int(chat_id), message_id=int(message_id), from_chat_id=from_chat_id)
                good_try += 1
            except BaseException:
                bad_try += 1
        await callback.message.reply(text=f'Рассылка совершена!\nУспешных отправок: {good_try}\n'
                                           f'Несостоявшихся отправок: {bad_try}')
    else:
        await callback.message.edit_text(text='Рассылка отменена')

@main_router.message(F.text != None)  # Хэндлер для обработки произвольных сообщений пользователя
async def answer_message(message: Message, bot: Bot):
    await message.answer(text=Lexicon_RU.get('answer_for_user'), reply_markup=await answer_for_user())


@main_router.message(F.web_app_data.data != None)  # Хэндлер для обработки заказа из webapp
async def web_app_order(message: Message, amo_api: AmoCRMWrapper, fields_id: dict, bot: Bot):
    raw_json = json.loads(message.web_app_data.data)
    logger.info(f'raw_json: {raw_json}')
    full_price = raw_json.get('total')
    contact_id = raw_json.get('userId')
    logger.info(f'contact_id: {contact_id}, full_price: {full_price}')
    custom_data = LeadData(raw_json=raw_json, fields_id=fields_id)
    try:

        if contact_id is None:
            raise ValueError('Нет id контакта к которому привязать заказ')

        order_data = Order(raw_json=raw_json, lead_id=111)
        if order_data.order_type == "commercial_offer":
            #Создание нового лида в статусе "КП отправлено"
            response = amo_api.send_lead_to_amo(pipeline_id=fields_id.get('pipeline_id'),
                                                status_id=fields_id.get('status_id_kp'),
                                                tags_data=custom_data.get_lead_tags(),
                                                contact_id=int(contact_id),
                                                price=int(full_price),
                                                custom_fields_data=custom_data.get_custom_fields_data())
            logger.info(f'Создана сделка {response}')
        else:
            # Создание нового заказа в статусе "Новый заказ"
            response = amo_api.send_lead_to_amo(pipeline_id=fields_id.get('pipeline_id'),
                                                status_id=fields_id.get('status_id_order'),
                                                tags_data=custom_data.get_lead_tags(),
                                                contact_id=int(contact_id),
                                                price=int(full_price),
                                                custom_fields_data=custom_data.get_custom_fields_data())
        lead_id = response.get('_embedded').get('leads')[0].get('id')
        logger.info(f'lead_id: {lead_id}')
        order_note = Order(raw_json=raw_json, lead_id=lead_id)

        # Добавление примечания в сделку
        amo_api.add_new_note_to_lead(lead_id=lead_id, text=order_note.get_order_message())

        # Добавление товаров в сделку
        items = raw_json.get('items')
        amo_api.add_catalog_elements_to_lead(lead_id=lead_id,
                                             catalog_id=fields_id.get('catalog_id'),
                                             elements=items)
        # Сообщения клиенту при запросе КП и отправка КП в чат бот.
        if order_data.order_type == 'commercial_offer':
            get_kp_pdf(lead_id)
            document = FSInputFile(f'Kp_{lead_id}.pdf')
            await message.answer_document(document=document, caption='КП по Вашему запросу!')
            os.remove(f'Kp_{lead_id}.pdf')
            await message.answer(text=Lexicon_RU.get('message_link_partners_kp'),
                                 reply_markup=await link_to_opt_button(lead_id=lead_id))

        # Сообщения клиенту при запросе счета
        else:
            # Первое сообщение клиенту о создании заказа
            await message.answer(text=order_note.get_order_message(service=False),
                                 reply_markup=ReplyKeyboardRemove())

            # Второе сообщение клиенту, приглашение в чат с партнёрами
            await message.answer(text=Lexicon_RU.get('message_link_partners_order'),
                                 reply_markup=await link_to_opt_button(lead_id=lead_id))

        # Отправка сообщения в чат проверки
        await bot.send_message(chat_id=fields_id.get('chat_id'),
                               text=f'Оформлен заказ:\n\n{order_note.get_order_message()}\n'
                                    f'Создана сделка: '
                                    f'<a href="https://hite.amocrm.ru/leads/detail/{lead_id}">{lead_id}</a>\n'
                                    f'Контакт клиента: '
                                    f'<a href="https://hite.amocrm.ru/contacts/detail/{contact_id}">{contact_id}</a>')

    except BaseException as error:
        logger.error(error)

        await message.answer(text='Произошла ошибка при отправке заказа, обратитесь к менеджеру.',
                             reply_markup=ReplyKeyboardRemove())

        await bot.send_message(chat_id=fields_id.get('chat_id'),
                               text=f'Произошла ошибка при оформлении заказа.\n\n{error}\n'
                                    f'Контакт клиента:'
                                    f'<a href="https://hite.amocrm.ru/contacts/detail/{contact_id}">{contact_id}</a>')






